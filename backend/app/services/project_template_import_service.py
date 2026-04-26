from __future__ import annotations

import hashlib
import mimetypes
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.exceptions import BadRequestException, NotFoundException, StorageException
from app.models.asset import Asset
from app.models.evaluation_item import EvaluationItem
from app.models.template import Template
from app.repositories.project_repository import ProjectRepository
from app.services.history_excel_import_service import HistoryExcelImportService


@dataclass(slots=True)
class ParsedProjectTemplateRow:
    sheet_name: str
    extension_type: str
    level2: str
    level3: str
    extension_standard: str | None
    control_point: str | None
    record_template: str | None
    default_compliance: str | None
    score_weight: float | None
    item_no: str | None
    source_row_no: int
    sort_order: int
    keywords_json: list[str]


class ProjectTemplateImportService(HistoryExcelImportService):
    TEMPLATE_TYPE = "project_record_reference"
    ASSET_KIND = "project_template_file"
    ASSET_CATEGORY = "template"
    ASSET_CATEGORY_LABEL = "结果记录参考模板"

    def __init__(self) -> None:
        super().__init__()
        self.project_repository = ProjectRepository()

    def import_excel(self, db: Session, project_id: str, filename: str, content: bytes) -> dict:
        project = self.project_repository.get(db, project_id)
        if not project:
            raise NotFoundException("PROJECT_NOT_FOUND", "项目不存在")

        workbook = self._load_workbook(filename, content)
        parsed_rows: list[ParsedProjectTemplateRow] = []
        skipped_count = 0
        matched_sheet_names: list[str] = []
        for sheet in workbook.worksheets:
            sheet_result = self._parse_sheet(sheet.title, sheet.iter_rows(values_only=True))
            if not sheet_result["matched"]:
                continue
            matched_sheet_names.append(sheet.title)
            parsed_rows.extend(sheet_result["rows"])
            skipped_count += sheet_result["skipped_count"]

        if not matched_sheet_names:
            raise BadRequestException("PROJECT_TEMPLATE_HEADER_NOT_FOUND", "未识别到模板表头，请确认包含扩展标准/控制点/测评项/结果记录/符合情况")
        if not parsed_rows:
            raise BadRequestException("PROJECT_TEMPLATE_EMPTY", "Excel 未解析出有效模板行")

        source_asset = self._store_source_asset(db, project_id, filename, content)
        self._deactivate_previous_templates(db, project_id)

        template = Template(
            project_id=project_id,
            name=Path(filename).stem,
            template_type=self.TEMPLATE_TYPE,
            extension_type=None,
            domain=None,
            level2=None,
            version=datetime.now().strftime("%Y%m%d%H%M%S"),
            description=f"来源文件：{filename}",
            source_asset_id=source_asset.id,
            is_builtin=False,
            is_active=True,
        )
        db.add(template)
        db.flush()

        items = [
            EvaluationItem(
                template_id=template.id,
                domain=None,
                level1=row.sheet_name,
                level2=row.level2,
                level3=row.level3,
                extension_type=row.extension_type,
                route_domain=None,
                source_template_name=filename,
                source_sheet_name=row.sheet_name,
                sort_order=row.sort_order,
                control_point=row.control_point,
                extension_standard=row.extension_standard,
                record_template=row.record_template,
                default_compliance=row.default_compliance,
                score_weight=row.score_weight,
                item_no=row.item_no,
                source_row_no=row.source_row_no,
                keywords_json=row.keywords_json,
                is_active=True,
            )
            for row in parsed_rows
        ]
        db.add_all(items)
        db.commit()
        db.refresh(template)

        return {
            "template_id": template.id,
            "source_asset_id": source_asset.id,
            "source_file": filename,
            "sheet_count": len(matched_sheet_names),
            "sheet_names": matched_sheet_names,
            "imported_count": len(items),
            "skipped_count": skipped_count,
        }

    def get_active_template_summary(self, db: Session, project_id: str) -> dict | None:
        if not self.project_repository.get(db, project_id):
            raise NotFoundException("PROJECT_NOT_FOUND", "项目不存在")

        template = (
            db.query(Template)
            .filter(
                Template.project_id == project_id,
                Template.template_type == self.TEMPLATE_TYPE,
                Template.is_active.is_(True),
            )
            .order_by(Template.created_at.desc())
            .first()
        )
        if not template:
            return None

        items = (
            db.query(EvaluationItem)
            .filter(EvaluationItem.template_id == template.id)
            .order_by(EvaluationItem.sort_order.asc(), EvaluationItem.created_at.asc())
            .all()
        )
        sheet_names = list(dict.fromkeys(item.source_sheet_name for item in items if item.source_sheet_name))
        source_file = template.source_asset.filename if template.source_asset else None
        return {
            "template_id": template.id,
            "project_id": template.project_id,
            "name": template.name,
            "template_type": template.template_type,
            "version": template.version,
            "source_asset_id": template.source_asset_id,
            "source_file": source_file,
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
            "item_count": len(items),
            "is_active": template.is_active,
            "created_at": template.created_at,
            "updated_at": template.updated_at,
        }

    def _parse_sheet(self, sheet_name: str, rows) -> dict:
        row_values = [list(row) for row in rows]
        header = self._find_header(row_values)
        if header is None:
            return {"matched": False, "rows": [], "skipped_count": 0}

        header_row_index, header_map = header
        parsed_rows: list[ParsedProjectTemplateRow] = []
        skipped_count = 0
        last_values: dict[str, str | None] = {key: None for key in self.HEADER_FALLBACK_FIELD_ORDER}
        extension_type = self._infer_asset_type(sheet_name) or "generic"
        for row_index in range(header_row_index + 2, len(row_values) + 1):
            row = row_values[row_index - 1]
            payload = self._build_row_payload(sheet_name, row_index, row, header_map, last_values, extension_type)
            if payload is None:
                skipped_count += 1
                continue
            parsed_rows.append(payload)
        return {"matched": True, "rows": parsed_rows, "skipped_count": skipped_count}

    def _build_row_payload(
        self,
        sheet_name: str,
        row_index: int,
        row: list[object],
        header_map: dict[str, int],
        last_values: dict[str, str | None],
        extension_type: str,
    ) -> ParsedProjectTemplateRow | None:
        values: dict[str, str | None] = {}
        for field_name, cell_index in header_map.items():
            raw = row[cell_index] if cell_index < len(row) else None
            normalized = self._normalize_value(raw)
            if normalized is not None:
                last_values[field_name] = normalized
            elif field_name in {"standard_type", "control_point"}:
                normalized = last_values[field_name]
            values[field_name] = normalized

        if not any(values.get(field) for field in ("control_point", "item_text", "raw_text", "item_code")):
            return None

        control_point = values.get("control_point")
        item_text = values.get("item_text")
        item_no = values.get("item_code")
        record_template = values.get("raw_text")
        level2 = control_point or sheet_name
        level3 = item_text or item_no or f"{sheet_name}-row-{row_index}"
        return ParsedProjectTemplateRow(
            sheet_name=sheet_name,
            extension_type=extension_type,
            level2=level2,
            level3=level3,
            extension_standard=values.get("standard_type"),
            control_point=control_point,
            record_template=record_template,
            default_compliance=values.get("compliance_result"),
            score_weight=self._parse_score(values.get("score_weight")),
            item_no=item_no,
            source_row_no=row_index,
            sort_order=row_index,
            keywords_json=self._extract_keywords(control_point, item_text, record_template),
        )

    def _load_workbook(self, filename: str, content: bytes):
        if not filename.lower().endswith(".xlsx"):
            raise BadRequestException("PROJECT_TEMPLATE_INVALID_TYPE", "仅支持上传 .xlsx 结果记录参考模板")
        try:
            return load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        except (InvalidFileException, OSError, ValueError) as exc:
            raise BadRequestException("PROJECT_TEMPLATE_INVALID_FILE", "结果记录参考模板解析失败") from exc

    def _store_source_asset(self, db: Session, project_id: str, filename: str, content: bytes) -> Asset:
        if not filename:
            raise BadRequestException("EMPTY_FILENAME", "上传文件名不能为空")

        upload_dir = Path(settings.UPLOAD_DIR) / project_id / "templates"
        upload_dir.mkdir(parents=True, exist_ok=True)

        original_name = Path(filename).name
        suffix = Path(original_name).suffix
        stored_name = f"{uuid4().hex}{suffix}"
        absolute_path = upload_dir / stored_name

        try:
            absolute_path.write_bytes(content)
        except OSError as exc:
            raise StorageException("FILE_UPLOAD_FAILED", "文件保存失败", str(exc)) from exc

        modified_at = datetime.fromtimestamp(absolute_path.stat().st_mtime)
        relative_path = str(Path("uploads") / project_id / "templates" / stored_name)
        batch_no = datetime.now().strftime("%Y%m%d%H%M%S")
        asset = Asset(
            project_id=project_id,
            asset_kind=self.ASSET_KIND,
            category=self.ASSET_CATEGORY,
            category_label=self.ASSET_CATEGORY_LABEL,
            batch_no=batch_no,
            filename=original_name,
            primary_ip=None,
            file_ext=suffix or None,
            mime_type=mimetypes.guess_type(original_name)[0],
            relative_path=relative_path,
            absolute_path=str(absolute_path),
            file_size=len(content),
            file_sha256=hashlib.sha256(content).hexdigest(),
            file_mtime=modified_at,
            source="upload",
            ingest_status="stored",
        )
        db.add(asset)
        db.flush()
        return asset

    def _deactivate_previous_templates(self, db: Session, project_id: str) -> None:
        templates = (
            db.query(Template)
            .filter(
                Template.project_id == project_id,
                Template.template_type == self.TEMPLATE_TYPE,
                Template.is_active.is_(True),
            )
            .all()
        )
        for template in templates:
            template.is_active = False
            db.add(template)

    def _extract_keywords(self, *parts: str | None) -> list[str]:
        return super()._extract_keywords(*parts)
