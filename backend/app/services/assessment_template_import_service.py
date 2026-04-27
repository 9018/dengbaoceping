from __future__ import annotations

import hashlib
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session

from app.core.exceptions import BadRequestException
from app.models.assessment_template import AssessmentTemplateItem, AssessmentTemplateSheet, AssessmentTemplateWorkbook
from app.models.knowledge_import_batch import KnowledgeImportBatch


@dataclass(slots=True)
class ParsedAssessmentTemplateItem:
    row_index: int
    standard_type: str | None
    control_point: str | None
    item_text: str | None
    record_template: str | None
    default_compliance_result: str | None
    weight: float | None
    item_code: str | None
    raw_row_json: dict
    page_types_json: list[str]
    required_facts_json: list[str]
    evidence_keywords_json: list[str]
    command_keywords_json: list[str]
    applicability_json: list[str]


@dataclass(slots=True)
class ParsedAssessmentTemplateSheet:
    sheet_name: str
    object_type: str | None
    object_category: str | None
    object_subtype: str | None
    is_physical: bool
    is_network: bool
    is_security_device: bool
    is_server: bool
    is_database: bool
    is_middleware: bool
    is_application: bool
    is_data_object: bool
    is_management: bool
    items: list[ParsedAssessmentTemplateItem]
    skipped_count: int


class AssessmentTemplateImportService:
    LIBRARY_TYPE = "assessment_template"
    ALLOWED_DUPLICATE_POLICIES = {"skip", "overwrite", "new_version"}
    REQUIRED_HEADERS = {
        "standard_type": ("扩展标准", "标准类型"),
        "control_point": ("控制点",),
        "item_text": ("测评项", "测评内容"),
        "record_template": ("结果记录", "测评记录"),
        "default_compliance_result": ("符合情况", "符合性结果"),
    }
    OPTIONAL_HEADERS = {
        "standard_type": ("扩展标准", "扩展标准:", "扩展标准：", "标准类型", "标准类型:", "标准类型："),
        "control_point": ("控制点", "控制点:", "控制点："),
        "item_text": ("测评项", "测评项:", "测评项：", "测评内容", "测评内容:", "测评内容："),
        "record_template": ("结果记录", "结果记录:", "结果记录：", "测评记录", "测评记录:", "测评记录："),
        "default_compliance_result": ("符合情况", "符合情况:", "符合情况：", "符合性结果", "符合性结果:", "符合性结果："),
        "weight": ("分值", "分值:", "分值：", "权重", "权重:", "权重："),
        "item_code": ("编号", "编号:", "编号：", "测评项编号", "测评项编号:", "测评项编号：", "条款编号", "条款编号:", "条款编号："),
    }
    HEADER_FIELDS = ("standard_type", "control_point", "item_text", "record_template", "default_compliance_result", "weight", "item_code")
    HEADER_STRIP_PATTERN = re.compile(r"[\s​‌‍⁠﻿]+")
    HEADER_TRAILING_PUNCTUATION_PATTERN = re.compile(r"[:：]+$")
    TOKEN_PATTERN = re.compile(r"[A-Za-z0-9_./-]{2,}|[一-鿿]{2,}")
    COMMAND_PATTERN = re.compile(r"\b(?:show|select|netstat|ipconfig|ifconfig|cat|grep|ps|wmic)\b", re.IGNORECASE)
    STOPWORDS = {"进行", "结果", "记录", "情况", "控制点", "测评项", "扩展标准", "符合", "部分符合", "不符合", "不适用"}
    OBJECT_RULES = (
        {"keywords": ("防火墙",), "object_type": "安全设备", "object_category": "防火墙", "object_subtype": None, "flags": {"is_security_device": True, "is_network": True}},
        {"keywords": ("交换机",), "object_type": "网络设备", "object_category": "交换机", "object_subtype": None, "flags": {"is_network": True}},
        {"keywords": ("windows",), "object_type": "服务器", "object_category": "Windows", "object_subtype": "Windows", "flags": {"is_server": True}},
        {"keywords": ("linux",), "object_type": "服务器", "object_category": "Linux", "object_subtype": "Linux", "flags": {"is_server": True}},
        {"keywords": ("mysql", "sqlserver", "sql server", "oracle", "数据库"), "object_type": "数据库", "object_category": "数据库", "object_subtype": None, "flags": {"is_database": True}},
        {"keywords": ("tomcat", "nginx", "iis", "中间件"), "object_type": "中间件", "object_category": "中间件", "object_subtype": None, "flags": {"is_middleware": True}},
        {"keywords": ("应用系统",), "object_type": "应用系统", "object_category": "应用系统", "object_subtype": None, "flags": {"is_application": True}},
        {"keywords": ("中心机房",), "object_type": "物理环境", "object_category": "中心机房", "object_subtype": None, "flags": {"is_physical": True}},
        {"keywords": ("安全通信网络",), "object_type": "通信网络", "object_category": "通信网络", "object_subtype": None, "flags": {"is_network": True}},
        {"keywords": ("系统边界",), "object_type": "区域边界", "object_category": "系统边界", "object_subtype": None, "flags": {"is_network": True}},
        {"keywords": ("重要业务数据", "个人信息数据"), "object_type": "数据对象", "object_category": "数据对象", "object_subtype": None, "flags": {"is_data_object": True}},
        {"keywords": ("安全管理",), "object_type": "管理制度", "object_category": "管理对象", "object_subtype": None, "flags": {"is_management": True}},
    )

    def import_excel(self, db: Session, filename: str, content: bytes, duplicate_policy: str = "skip") -> dict:
        policy = self._normalize_duplicate_policy(duplicate_policy)
        source_file_hash = self._build_source_file_hash(content)
        workbook = self._load_workbook(filename, content)
        workbook_name = self._build_workbook_name(filename)
        version = self._extract_version(filename)
        parsed_sheets: list[ParsedAssessmentTemplateSheet] = []
        skipped_count = 0
        for worksheet in workbook.worksheets:
            parsed = self._parse_sheet(worksheet.title, worksheet.iter_rows(values_only=True))
            if parsed is None:
                continue
            parsed_sheets.append(parsed)
            skipped_count += parsed.skipped_count
        if not parsed_sheets:
            raise BadRequestException("ASSESSMENT_TEMPLATE_HEADER_NOT_FOUND", "未识别到结果记录模板表头，请确认包含扩展标准/控制点/测评项/结果记录/符合情况")
        item_count = sum(len(sheet.items) for sheet in parsed_sheets)
        if item_count == 0:
            raise BadRequestException("ASSESSMENT_TEMPLATE_EMPTY", "Excel 未解析出有效模板项")

        duplicate_workbook = self._find_duplicate_workbook(db, source_file_hash)
        batch = KnowledgeImportBatch(
            library_type=self.LIBRARY_TYPE,
            source_file=filename,
            source_file_hash=source_file_hash,
            file_size=len(content),
            item_count=item_count,
            status="pending",
            duplicate_of_id=duplicate_workbook.import_batch_id if duplicate_workbook else None,
            import_mode=policy,
        )
        db.add(batch)
        db.flush()

        sheet_names = [parsed_sheet.sheet_name for parsed_sheet in parsed_sheets]
        duplicate = duplicate_workbook is not None
        if duplicate and policy == "skip":
            batch.status = "skipped"
            batch.summary_json = self._build_batch_summary(
                sheet_count=len(parsed_sheets),
                sheet_names=sheet_names,
                item_count=item_count,
                imported_count=0,
                skipped_count=item_count,
            )
            db.commit()
            return {
                "workbook_id": duplicate_workbook.id,
                "batch_id": batch.id,
                "source_file": filename,
                "source_file_hash": source_file_hash,
                "name": workbook_name,
                "version": version,
                "sheet_count": len(parsed_sheets),
                "sheet_names": sheet_names,
                "item_count": item_count,
                "imported_count": 0,
                "skipped_count": item_count,
                "duplicate": True,
                "duplicate_policy": policy,
                "duplicate_of_id": batch.duplicate_of_id,
                "status": batch.status,
            }

        if policy == "overwrite":
            self._archive_active_workbooks(db, filename)

        workbook_record = AssessmentTemplateWorkbook(
            source_file=filename,
            source_file_hash=source_file_hash,
            file_size=len(content),
            import_batch_id=batch.id,
            is_archived=False,
            name=workbook_name,
            version=version,
            sheet_count=len(parsed_sheets),
            item_count=item_count,
        )
        db.add(workbook_record)
        db.flush()

        for parsed_sheet in parsed_sheets:
            sheet_record = AssessmentTemplateSheet(
                workbook_id=workbook_record.id,
                sheet_name=parsed_sheet.sheet_name,
                object_type=parsed_sheet.object_type,
                object_category=parsed_sheet.object_category,
                object_subtype=parsed_sheet.object_subtype,
                is_physical=parsed_sheet.is_physical,
                is_network=parsed_sheet.is_network,
                is_security_device=parsed_sheet.is_security_device,
                is_server=parsed_sheet.is_server,
                is_database=parsed_sheet.is_database,
                is_middleware=parsed_sheet.is_middleware,
                is_application=parsed_sheet.is_application,
                is_data_object=parsed_sheet.is_data_object,
                is_management=parsed_sheet.is_management,
                row_count=len(parsed_sheet.items),
            )
            db.add(sheet_record)
            db.flush()
            item_records = [
                AssessmentTemplateItem(
                    workbook_id=workbook_record.id,
                    sheet_id=sheet_record.id,
                    sheet_name=parsed_sheet.sheet_name,
                    row_index=item.row_index,
                    standard_type=item.standard_type,
                    control_point=item.control_point,
                    item_text=item.item_text,
                    record_template=item.record_template,
                    default_compliance_result=item.default_compliance_result,
                    weight=item.weight,
                    item_code=item.item_code,
                    object_type=parsed_sheet.object_type,
                    object_category=parsed_sheet.object_category,
                    page_types_json=item.page_types_json,
                    required_facts_json=item.required_facts_json,
                    evidence_keywords_json=item.evidence_keywords_json,
                    command_keywords_json=item.command_keywords_json,
                    applicability_json=item.applicability_json,
                    raw_row_json=item.raw_row_json,
                )
                for item in parsed_sheet.items
            ]
            db.add_all(item_records)

        batch.status = "imported"
        batch.summary_json = self._build_batch_summary(
            sheet_count=len(parsed_sheets),
            sheet_names=sheet_names,
            item_count=item_count,
            imported_count=item_count,
            skipped_count=skipped_count,
        )
        db.commit()
        return {
            "workbook_id": workbook_record.id,
            "batch_id": batch.id,
            "source_file": filename,
            "source_file_hash": source_file_hash,
            "name": workbook_record.name,
            "version": workbook_record.version,
            "sheet_count": workbook_record.sheet_count,
            "sheet_names": sheet_names,
            "item_count": workbook_record.item_count,
            "imported_count": workbook_record.item_count,
            "skipped_count": skipped_count,
            "duplicate": duplicate,
            "duplicate_policy": policy,
            "duplicate_of_id": batch.duplicate_of_id,
            "status": batch.status,
        }

    def _normalize_duplicate_policy(self, duplicate_policy: str) -> str:
        normalized = (duplicate_policy or "skip").strip().lower()
        if normalized not in self.ALLOWED_DUPLICATE_POLICIES:
            raise BadRequestException(
                "ASSESSMENT_TEMPLATE_DUPLICATE_POLICY_INVALID",
                "duplicate_policy 仅支持 skip、overwrite、new_version",
            )
        return normalized

    def _build_source_file_hash(self, content: bytes) -> str:
        return hashlib.sha256(content).hexdigest()

    def _find_duplicate_workbook(self, db: Session, source_file_hash: str) -> AssessmentTemplateWorkbook | None:
        return (
            db.query(AssessmentTemplateWorkbook)
            .filter(AssessmentTemplateWorkbook.source_file_hash == source_file_hash)
            .order_by(AssessmentTemplateWorkbook.created_at.desc())
            .first()
        )

    def _archive_active_workbooks(self, db: Session, source_file: str) -> None:
        (
            db.query(AssessmentTemplateWorkbook)
            .filter(
                AssessmentTemplateWorkbook.source_file == source_file,
                AssessmentTemplateWorkbook.is_archived.is_(False),
            )
            .update({AssessmentTemplateWorkbook.is_archived: True}, synchronize_session=False)
        )

    def _build_batch_summary(
        self,
        *,
        sheet_count: int,
        sheet_names: list[str],
        item_count: int,
        imported_count: int,
        skipped_count: int,
    ) -> dict:
        return {
            "sheet_count": sheet_count,
            "sheet_names": sheet_names,
            "item_count": item_count,
            "imported_count": imported_count,
            "skipped_count": skipped_count,
        }

    def _load_workbook(self, filename: str, content: bytes):
        if not filename.lower().endswith(".xlsx"):
            raise BadRequestException("ASSESSMENT_TEMPLATE_INVALID_TYPE", "仅支持上传 .xlsx 结果记录模板文件")
        try:
            return load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        except (InvalidFileException, OSError, ValueError) as exc:
            raise BadRequestException("ASSESSMENT_TEMPLATE_INVALID_FILE", "结果记录模板 Excel 解析失败") from exc

    def _parse_sheet(self, sheet_name: str, rows) -> ParsedAssessmentTemplateSheet | None:
        row_values = [list(row) for row in rows]
        header = self._find_header(row_values)
        if header is None:
            return None
        header_row_index, header_map = header
        items, skipped_count = self._parse_sheet_rows(row_values, header_row_index, header_map)
        profile = self._infer_sheet_object_profile(sheet_name)
        return ParsedAssessmentTemplateSheet(
            sheet_name=sheet_name,
            object_type=profile["object_type"],
            object_category=profile["object_category"],
            object_subtype=profile["object_subtype"],
            is_physical=profile["is_physical"],
            is_network=profile["is_network"],
            is_security_device=profile["is_security_device"],
            is_server=profile["is_server"],
            is_database=profile["is_database"],
            is_middleware=profile["is_middleware"],
            is_application=profile["is_application"],
            is_data_object=profile["is_data_object"],
            is_management=profile["is_management"],
            items=items,
            skipped_count=skipped_count,
        )

    def _find_header(self, rows: list[list[object]]) -> tuple[int, dict[str, int]] | None:
        if len(rows) >= 2:
            header_map = self._match_header_row(rows[1])
            if header_map is not None:
                return 1, header_map
        for index, row in enumerate(rows):
            header_map = self._match_header_row(row)
            if header_map is not None:
                return index, header_map
        return None

    def _match_header_row(self, row: list[object]) -> dict[str, int] | None:
        normalized_cells = [self._normalize_cell(cell) for cell in row]
        aliases_map = {
            field_name: {self._normalize_cell(alias) for alias in aliases}
            for field_name, aliases in self.OPTIONAL_HEADERS.items()
        }
        header_map: dict[str, int] = {}
        for field_name, aliases in aliases_map.items():
            for cell_index, cell in enumerate(normalized_cells):
                if cell in aliases:
                    header_map[field_name] = cell_index
                    break
        if all(field in header_map for field in self.REQUIRED_HEADERS):
            return header_map
        return None

    def _parse_sheet_rows(self, rows: list[list[object]], header_row_index: int, header_map: dict[str, int]) -> tuple[list[ParsedAssessmentTemplateItem], int]:
        parsed_items: list[ParsedAssessmentTemplateItem] = []
        skipped_count = 0
        last_values: dict[str, str | None] = {"standard_type": None, "control_point": None}
        for row_index in range(header_row_index + 2, len(rows) + 1):
            row = rows[row_index - 1]
            payload = self._build_item_payload(row_index, row, header_map, last_values)
            if payload is None:
                skipped_count += 1
                continue
            parsed_items.append(payload)
        return parsed_items, skipped_count

    def _build_item_payload(
        self,
        row_index: int,
        row: list[object],
        header_map: dict[str, int],
        last_values: dict[str, str | None],
    ) -> ParsedAssessmentTemplateItem | None:
        values: dict[str, str | None] = {}
        cells = [self._normalize_value(cell) for cell in row]
        for field_name in self.HEADER_FIELDS:
            cell_index = header_map.get(field_name)
            raw = row[cell_index] if cell_index is not None and cell_index < len(row) else None
            normalized = self._normalize_value(raw)
            if normalized is not None and field_name in {"standard_type", "control_point"}:
                last_values[field_name] = normalized
            elif normalized is None and field_name in {"standard_type", "control_point"}:
                normalized = last_values[field_name]
            values[field_name] = normalized
        if not values.get("item_text") and not values.get("record_template"):
            return None
        metadata = self._build_item_metadata(values.get("control_point"), values.get("item_text"), values.get("record_template"))
        return ParsedAssessmentTemplateItem(
            row_index=row_index,
            standard_type=values.get("standard_type"),
            control_point=values.get("control_point"),
            item_text=values.get("item_text"),
            record_template=values.get("record_template"),
            default_compliance_result=values.get("default_compliance_result"),
            weight=self._parse_score(values.get("weight")),
            item_code=values.get("item_code"),
            raw_row_json={
                "cells": cells,
                "alternative_templates": self._extract_alternative_templates(row, header_map),
            },
            page_types_json=metadata["page_types_json"],
            required_facts_json=metadata["required_facts_json"],
            evidence_keywords_json=metadata["evidence_keywords_json"],
            command_keywords_json=metadata["command_keywords_json"],
            applicability_json=metadata["applicability_json"],
        )

    def _extract_alternative_templates(self, row: list[object], header_map: dict[str, int]) -> list[str]:
        last_header_index = max(header_map.values(), default=-1)
        alternatives: list[str] = []
        for cell in row[last_header_index + 1 : last_header_index + 4]:
            normalized = self._normalize_value(cell)
            if normalized:
                alternatives.append(normalized)
        return alternatives

    def _build_item_metadata(self, control_point: str | None, item_text: str | None, record_template: str | None) -> dict[str, list[str]]:
        text = " ".join(part for part in (control_point, item_text, record_template) if part)
        tokens = self._extract_keywords(text)
        page_types: list[str] = []
        if any(token in text.lower() for token in ("身份鉴别", "口令", "密码", "登录")):
            page_types.append("password_policy")
        if any(token in text.lower() for token in ("访问控制", "acl", "权限")):
            page_types.append("access_control_policy")
        if any(token in text.lower() for token in ("安全审计", "审计", "日志")):
            page_types.append("audit_log")
        if any(token in text.lower() for token in ("入侵", "病毒", "恶意代码")):
            page_types.append("intrusion_prevention")
        command_keywords = sorted({match.group(0).lower() for match in self.COMMAND_PATTERN.finditer(text)})
        required_facts = [token for token in tokens if token not in command_keywords][:8]
        applicability = [token for token in tokens if token in {"windows", "linux", "防火墙", "交换机", "数据库", "中间件", "应用系统"}]
        return {
            "page_types_json": page_types,
            "required_facts_json": required_facts,
            "evidence_keywords_json": tokens[:12],
            "command_keywords_json": command_keywords,
            "applicability_json": applicability,
        }

    def _infer_sheet_object_profile(self, sheet_name: str) -> dict[str, str | bool | None]:
        normalized = unicodedata.normalize("NFKC", sheet_name).lower().replace(" ", "")
        profile = {
            "object_type": None,
            "object_category": None,
            "object_subtype": None,
            "is_physical": False,
            "is_network": False,
            "is_security_device": False,
            "is_server": False,
            "is_database": False,
            "is_middleware": False,
            "is_application": False,
            "is_data_object": False,
            "is_management": False,
        }
        for rule in self.OBJECT_RULES:
            if any(keyword.lower().replace(" ", "") in normalized for keyword in rule["keywords"]):
                profile["object_type"] = rule["object_type"]
                profile["object_category"] = rule["object_category"]
                profile["object_subtype"] = rule["object_subtype"]
                for key, value in rule["flags"].items():
                    profile[key] = value
                break
        return profile

    def _build_workbook_name(self, filename: str) -> str:
        normalized = filename.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
        return normalized.rsplit(".", 1)[0]

    def _extract_version(self, filename: str) -> str | None:
        match = re.search(r"(20\d{6})", filename)
        return match.group(1) if match else None

    def _normalize_cell(self, cell: object) -> str:
        text = unicodedata.normalize("NFKC", str(cell or ""))
        text = self.HEADER_STRIP_PATTERN.sub("", text)
        text = self.HEADER_TRAILING_PUNCTUATION_PATTERN.sub("", text)
        return text.strip()

    def _normalize_value(self, value: object) -> str | None:
        if value is None:
            return None
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        text = str(value).strip()
        return text or None

    def _parse_score(self, value: str | None) -> float | None:
        if value is None:
            return None
        try:
            return float(value)
        except ValueError:
            return None

    def _extract_keywords(self, text: str) -> list[str]:
        counter: Counter[str] = Counter()
        for token in self.TOKEN_PATTERN.findall(text):
            normalized = token.strip().lower()
            if len(normalized) < 2 or normalized in self.STOPWORDS:
                continue
            counter[normalized] += 1
        return [item for item, _ in counter.most_common(16)]
