from __future__ import annotations

import re
import unicodedata
from collections import Counter
from dataclasses import asdict, dataclass
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session

from app.core.exceptions import BadRequestException
from app.models.history_record import HistoryRecord
from app.repositories.history_record_repository import HistoryRecordRepository


@dataclass(slots=True)
class ParsedHistoryRow:
    source_file: str
    sheet_name: str
    asset_name: str
    asset_type: str | None
    extension_standard: str | None
    control_point: str | None
    evaluation_item: str | None
    record_text: str | None
    compliance_status: str | None
    score: float | None
    item_no: str | None
    row_index: int
    keywords_json: list[str]


class HistoryImportService:
    REQUIRED_HEADERS = {
        "extension_standard": ("扩展标准",),
        "control_point": ("控制点",),
        "evaluation_item": ("测评项",),
        "record_text": ("结果记录",),
        "compliance_status": ("符合情况",),
        "score": ("分值",),
        "item_no": ("编号",),
    }
    OPTIONAL_HEADERS = {
        "extension_standard": ("扩展标准", "扩展标准:", "扩展标准："),
        "control_point": ("控制点", "控制点:", "控制点："),
        "evaluation_item": ("测评项", "测评项:", "测评项："),
        "record_text": ("结果记录", "结果记录:", "结果记录："),
        "compliance_status": ("符合情况", "符合情况:", "符合情况："),
        "score": ("分值", "分值:", "分值："),
        "item_no": ("编号", "编号:", "编号："),
    }
    HEADER_STRIP_PATTERN = re.compile(r"[\s​‌‍⁠﻿]+")
    HEADER_TRAILING_PUNCTUATION_PATTERN = re.compile(r"[:：]+$")
    HEADER_FALLBACK_FIELD_ORDER = (
        "extension_standard",
        "control_point",
        "evaluation_item",
        "record_text",
        "compliance_status",
        "score",
        "item_no",
    )
    ASSET_TYPE_RULES = (
        ("防火墙", "firewall"),
        ("交换机", "switch"),
        ("服务器", "server"),
        ("数据库", "database"),
        ("安全管理", "management"),
    )
    TOKEN_PATTERN = re.compile(r"[A-Za-z0-9_-]{3,}|[一-鿿]{2,}")
    STOPWORDS = {"进行", "结果", "记录", "情况", "控制点", "测评项", "扩展标准", "符合", "部分符合", "不符合", "不适用"}

    def __init__(self) -> None:
        self.repository = HistoryRecordRepository()

    def import_excel(self, db: Session, filename: str, content: bytes) -> dict:
        workbook = self._load_workbook(filename, content)
        parsed_rows: list[ParsedHistoryRow] = []
        skipped_count = 0
        matched_sheet_count = 0
        for sheet in workbook.worksheets:
            sheet_result = self._parse_sheet(sheet.title, sheet.iter_rows(values_only=True), filename)
            if not sheet_result["matched"]:
                continue
            matched_sheet_count += 1
            parsed_rows.extend(sheet_result["rows"])
            skipped_count += sheet_result["skipped_count"]
        if matched_sheet_count == 0:
            raise BadRequestException("HISTORY_EXCEL_HEADER_NOT_FOUND", "未识别到历史测评记录表头，请确认包含扩展标准/控制点/测评项/结果记录/符合情况/分值/编号")
        if not parsed_rows:
            raise BadRequestException("HISTORY_EXCEL_EMPTY", "Excel 未解析出有效历史记录")
        records = [HistoryRecord(**asdict(row)) for row in parsed_rows]
        self.repository.create_many(db, records)
        db.commit()
        return {
            "source_file": filename,
            "sheet_count": len(workbook.sheetnames),
            "imported_count": len(records),
            "skipped_count": skipped_count,
            "compliance_status_counts": self._build_status_counts(records),
        }

    def _load_workbook(self, filename: str, content: bytes):
        if not filename.lower().endswith(".xlsx"):
            raise BadRequestException("HISTORY_EXCEL_INVALID_TYPE", "仅支持上传 .xlsx 历史测评记录文件")
        try:
            return load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        except (InvalidFileException, OSError, ValueError) as exc:
            raise BadRequestException("HISTORY_EXCEL_INVALID_FILE", "历史测评记录 Excel 解析失败") from exc

    def _parse_sheet(self, sheet_name: str, rows, source_file: str) -> dict:
        row_values = [list(row) for row in rows]
        header = self._find_header(row_values)
        if header is None:
            return {"matched": False, "rows": [], "skipped_count": 0}
        header_row_index, header_map = header
        parsed_rows: list[ParsedHistoryRow] = []
        skipped_count = 0
        last_values: dict[str, str | None] = {key: None for key in self.REQUIRED_HEADERS}
        for row_index in range(header_row_index + 2, len(row_values) + 1):
            row = row_values[row_index - 1]
            payload = self._build_row_payload(sheet_name, source_file, row_index, row, header_map, last_values)
            if payload is None:
                skipped_count += 1
                continue
            parsed_rows.append(payload)
        return {"matched": True, "rows": parsed_rows, "skipped_count": skipped_count}

    def _find_header(self, rows: list[list[object]]) -> tuple[int, dict[str, int]] | None:
        normalized_required_headers = {
            field_name: {self._normalize_cell(alias) for alias in aliases}
            for field_name, aliases in self.OPTIONAL_HEADERS.items()
        }
        for index, row in enumerate(rows):
            normalized_cells = [self._normalize_cell(cell) for cell in row]
            header_map: dict[str, int] = {}
            for field_name, aliases in normalized_required_headers.items():
                for cell_index, cell in enumerate(normalized_cells):
                    if cell in aliases:
                        header_map[field_name] = cell_index
                        break
            if len(header_map) == len(self.REQUIRED_HEADERS):
                return index, header_map
            fallback_map = self._build_fallback_header_map(normalized_cells, header_map)
            if fallback_map is not None:
                return index, fallback_map
        return None

    def _build_fallback_header_map(self, normalized_cells: list[str], header_map: dict[str, int]) -> dict[str, int] | None:
        if not header_map:
            return None
        first_missing_field = next((field for field in self.HEADER_FALLBACK_FIELD_ORDER if field not in header_map), None)
        if first_missing_field is None:
            return header_map
        first_missing_index = self.HEADER_FALLBACK_FIELD_ORDER.index(first_missing_field)
        if any(field not in header_map for field in self.HEADER_FALLBACK_FIELD_ORDER[:first_missing_index]):
            return None
        fallback_map = dict(header_map)
        last_index = -1
        for field_name in self.HEADER_FALLBACK_FIELD_ORDER:
            if field_name in fallback_map:
                last_index = fallback_map[field_name]
                continue
            last_index += 1
            if last_index >= len(normalized_cells):
                return None
            remaining_fields = self.HEADER_FALLBACK_FIELD_ORDER[self.HEADER_FALLBACK_FIELD_ORDER.index(field_name) + 1 :]
            if any(next_field in fallback_map for next_field in remaining_fields):
                return None
            fallback_map[field_name] = last_index
        if not any(self._normalize_cell(cell) for cell in normalized_cells):
            return None
        return fallback_map

    def _build_row_payload(
        self,
        sheet_name: str,
        source_file: str,
        row_index: int,
        row: list[object],
        header_map: dict[str, int],
        last_values: dict[str, str | None],
    ) -> ParsedHistoryRow | None:
        values: dict[str, str | None] = {}
        for field_name, cell_index in header_map.items():
            raw = row[cell_index] if cell_index < len(row) else None
            normalized = self._normalize_value(raw)
            if normalized is not None:
                last_values[field_name] = normalized
            elif field_name in {"extension_standard", "control_point", "evaluation_item", "item_no"}:
                normalized = last_values[field_name]
            values[field_name] = normalized
        if not any(values.get(field) for field in ("control_point", "evaluation_item", "record_text", "item_no")):
            return None
        return ParsedHistoryRow(
            source_file=source_file,
            sheet_name=sheet_name,
            asset_name=sheet_name.strip() or sheet_name,
            asset_type=self._infer_asset_type(sheet_name),
            extension_standard=values.get("extension_standard"),
            control_point=values.get("control_point"),
            evaluation_item=values.get("evaluation_item"),
            record_text=values.get("record_text"),
            compliance_status=values.get("compliance_status"),
            score=self._parse_score(values.get("score")),
            item_no=values.get("item_no"),
            row_index=row_index,
            keywords_json=self._extract_keywords(values.get("control_point"), values.get("evaluation_item"), values.get("record_text")),
        )

    def _normalize_cell(self, cell: object) -> str:
        text = unicodedata.normalize("NFKC", str(cell or ""))
        text = self.HEADER_STRIP_PATTERN.sub("", text)
        text = self.HEADER_TRAILING_PUNCTUATION_PATTERN.sub("", text)
        return text.strip()

    def _normalize_value(self, value: object) -> str | None:
        if value is None:
            return None
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        text = str(value).strip()
        return text or None

    def _parse_score(self, value: str | None) -> float | None:
        if value is None:
            return None
        try:
            return float(value)
        except ValueError:
            return None

    def _infer_asset_type(self, sheet_name: str) -> str | None:
        normalized = sheet_name.strip()
        for keyword, asset_type in self.ASSET_TYPE_RULES:
            if keyword in normalized:
                return asset_type
        return None

    def _extract_keywords(self, *parts: str | None) -> list[str]:
        counter: Counter[str] = Counter()
        for token in self.TOKEN_PATTERN.findall(" ".join(part or "" for part in parts)):
            normalized = token.strip().lower()
            if len(normalized) < 2 or normalized in self.STOPWORDS:
                continue
            counter[normalized] += 1
        return [item for item, _ in counter.most_common(12)]

    def _build_status_counts(self, records: list[HistoryRecord]) -> dict[str, int]:
        counts: dict[str, int] = {}
        for record in records:
            key = record.compliance_status or "未标注"
            counts[key] = counts.get(key, 0) + 1
        return counts
