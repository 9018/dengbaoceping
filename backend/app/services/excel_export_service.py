from __future__ import annotations

from collections import defaultdict
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from app.models.evaluation_record import EvaluationRecord
from app.repositories.guidance_history_link_repository import GuidanceHistoryLinkRepository

OFFICIAL_HEADERS = ["编号", "扩展标准", "控制点", "测评项", "结果记录", "符合情况", "分值"]
DEBUG_HEADERS = OFFICIAL_HEADERS + ["测试对象", "证据文件", "指导书依据", "匹配分数", "历史样本ID"]
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E8FB")
HEADER_FONT = Font(bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_SIDE = Side(style="thin", color="D0D7DE")
HEADER_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
SHEET_NAME_MAX_LENGTH = 31


class ExcelExportService:
    def __init__(self) -> None:
        self.guidance_history_repository = GuidanceHistoryLinkRepository()

    def build_workbook(self, db: Session, records: list[EvaluationRecord], mode: str) -> bytes:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        export_rows = self._build_export_rows(db, records)
        grouped_rows: dict[str, list[dict]] = defaultdict(list)
        for row in export_rows:
            grouped_rows[row["sheet_name"]].append(row)

        existing_names: set[str] = set()
        for original_sheet_name, rows in grouped_rows.items():
            sheet_name = self._dedupe_sheet_name(original_sheet_name, existing_names)
            worksheet = workbook.create_sheet(sheet_name)
            headers = DEBUG_HEADERS if mode == "debug" else OFFICIAL_HEADERS
            worksheet.append(headers)
            self._style_header(worksheet)
            for row in rows:
                values = [
                    row["item_no"],
                    row["extension_standard"],
                    row["control_point"],
                    row["evaluation_item"],
                    row["result_record"],
                    row["compliance_status"],
                    row["score"],
                ]
                if mode == "debug":
                    values.extend(
                        [
                            row["asset_name"],
                            row["evidence_files"],
                            row["guidance_basis"],
                            row["match_score_text"],
                            row["history_sample_id"],
                        ]
                    )
                worksheet.append(values)
            worksheet.freeze_panes = "A2"
            self._auto_fit_columns(worksheet)

        if not workbook.sheetnames:
            worksheet = workbook.create_sheet("未命名资产")
            worksheet.append(DEBUG_HEADERS if mode == "debug" else OFFICIAL_HEADERS)
            self._style_header(worksheet)
            worksheet.freeze_panes = "A2"
            self._auto_fit_columns(worksheet)

        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def _build_export_rows(self, db: Session, records: list[EvaluationRecord]) -> list[dict]:
        guidance_ids = self._collect_guidance_ids(records)
        history_rows = self.guidance_history_repository.list_history_by_guidance_ids(db, guidance_ids)
        history_by_guidance: dict[str, tuple] = {}
        for link, history in history_rows:
            history_by_guidance.setdefault(link.guidance_item_id, (link, history))

        rows = []
        for record in reversed(records):
            evidence_items = [link.evidence for link in getattr(record, "evidence_links", []) if getattr(link, "evidence", None)]
            primary_evidence = evidence_items[0] if evidence_items else None
            matched_guidance = next((item.matched_guidance for item in evidence_items if getattr(item, "matched_guidance", None)), None)
            guidance_id = matched_guidance.id if matched_guidance else None
            history_link, history_record = history_by_guidance.get(guidance_id, (None, None))
            asset_name = self._resolve_asset_name(record, evidence_items)
            rows.append(
                {
                    "sheet_name": asset_name,
                    "item_no": self._coalesce(history_record.item_no if history_record else None, record.record_no),
                    "extension_standard": self._coalesce(history_record.extension_standard if history_record else None, record.indicator_l2),
                    "control_point": self._coalesce(history_record.control_point if history_record else None, record.indicator_l3, record.title),
                    "evaluation_item": self._coalesce(record.title, getattr(record.evaluation_item, "level3", None), history_record.evaluation_item if history_record else None),
                    "result_record": self._coalesce(record.final_content, record.record_text),
                    "compliance_status": self._resolve_compliance_status(record, history_record),
                    "score": history_record.score if history_record else None,
                    "asset_name": asset_name,
                    "evidence_files": "\n".join(filter(None, [item.title for item in evidence_items])) or None,
                    "guidance_basis": self._build_guidance_basis(matched_guidance),
                    "match_score_text": self._build_match_score_text(record, primary_evidence, history_link),
                    "history_sample_id": history_record.id if history_record else None,
                }
            )
        return rows

    def _collect_guidance_ids(self, records: list[EvaluationRecord]) -> list[str]:
        guidance_ids: list[str] = []
        for record in records:
            for link in getattr(record, "evidence_links", []):
                evidence = getattr(link, "evidence", None)
                matched_guidance = getattr(evidence, "matched_guidance", None)
                if matched_guidance and matched_guidance.id not in guidance_ids:
                    guidance_ids.append(matched_guidance.id)
        return guidance_ids

    def _resolve_asset_name(self, record: EvaluationRecord, evidence_items: list) -> str:
        for evidence in evidence_items:
            matched_asset = getattr(evidence, "matched_asset", None)
            if matched_asset and matched_asset.filename:
                return matched_asset.filename
        asset = getattr(record, "asset", None)
        if asset and asset.filename:
            return asset.filename
        for evidence in evidence_items:
            if evidence.device:
                return evidence.device
        return "未命名资产"

    def _build_guidance_basis(self, guidance_item) -> str | None:
        if not guidance_item:
            return None
        code = guidance_item.guidance_code or ""
        title = guidance_item.section_title or ""
        path = guidance_item.section_path or ""
        parts = [part for part in [code, title] if part]
        base = " - ".join(parts)
        if path:
            return f"{base} ({path})" if base else path
        return base or None

    def _build_match_score_text(self, record: EvaluationRecord, evidence, history_link) -> str | None:
        parts: list[str] = []
        if record.match_score is not None:
            parts.append(f"记录:{record.match_score}")
        if evidence and evidence.guidance_match_score is not None:
            parts.append(f"指导书:{evidence.guidance_match_score}")
        if history_link is not None:
            parts.append(f"历史:{history_link.match_score}")
        return " / ".join(parts) or None

    def _resolve_compliance_status(self, record: EvaluationRecord, history_record) -> str | None:
        reasons = record.match_reasons_json if isinstance(record.match_reasons_json, dict) else {}
        generation = reasons.get("record_generation") if isinstance(reasons.get("record_generation"), dict) else {}
        return self._coalesce(record.conclusion, generation.get("compliance_result"), history_record.compliance_status if history_record else None)

    def _style_header(self, worksheet) -> None:
        for cell in worksheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = HEADER_BORDER

    def _auto_fit_columns(self, worksheet) -> None:
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 50)

    def _dedupe_sheet_name(self, raw_name: str, existing_names: set[str]) -> str:
        sanitized = self._sanitize_sheet_name(raw_name)
        candidate = sanitized
        suffix = 1
        while candidate in existing_names:
            token = f"_{suffix}"
            candidate = f"{sanitized[: SHEET_NAME_MAX_LENGTH - len(token)]}{token}"
            suffix += 1
        existing_names.add(candidate)
        return candidate

    def _sanitize_sheet_name(self, raw_name: str | None) -> str:
        value = (raw_name or "未命名资产").strip()
        for char in "[]:*?/\\":
            value = value.replace(char, "-")
        value = value[:SHEET_NAME_MAX_LENGTH].strip()
        return value or "未命名资产"

    def _coalesce(self, *values):
        for value in values:
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            return value
        return None
