from io import BytesIO

from openpyxl import Workbook


STANDARD_HEADERS = ["扩展标准", "控制点", "测评项", "结果记录", "符合情况", "分值", "编号"]


def build_history_excel() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "出口防火墙"
    worksheet.append(STANDARD_HEADERS)
    worksheet.append(["安全通信网络", "边界访问控制", "应限制非授权访问", "经现场核查，已配置访问控制策略。", "符合", 1.0, "A-01"])
    worksheet.append([None, "边界访问控制", "应保留访问控制日志", "查看日志审计配置，当前已开启日志留存。", "部分符合", 0.6, "A-02"])

    second = workbook.create_sheet("核心交换机")
    second.append(STANDARD_HEADERS)
    second.append(["安全区域边界", "网络访问控制", "应配置 ACL", "未提供完整 ACL 配置截图。", "不符合", 0.2, "B-01"])
    second.append(["安全区域边界", "网络访问控制", "应定期核查", "当前设备为汇聚交换机，不适用该项。", "不适用", 0.0, "B-02"])

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_history_excel_with_variation_headers() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "出口防火墙"
    worksheet.append(["﻿扩展标准", "控制点\n", "测评项", "结果记录：", "符合情况", "分值", "编号"])
    worksheet.append(["安全通信网络", "边界访问控制", "应限制非授权访问", "经现场核查，已配置访问控制策略。", "符合", 1.0, "A-01"])

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_history_excel_with_cover_sheet() -> bytes:
    workbook = Workbook()
    cover = workbook.active
    cover.title = "封面"
    cover.append(["历史测评记录导入模板"])
    cover.append(["说明", "本页非数据页"])

    worksheet = workbook.create_sheet("出口防火墙")
    worksheet.append(STANDARD_HEADERS)
    worksheet.append(["安全通信网络", "边界访问控制", "应限制非授权访问", "经现场核查，已配置访问控制策略。", "符合", 1.0, "A-01"])

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_history_excel_with_trailing_blank_headers() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "出口防火墙"
    worksheet.append(["扩展标准", "控制点 ", "测评项", "结果记录", "符合情况", None, None])
    worksheet.append(["安全通信网络", "边界访问控制", "应限制非授权访问", "经现场核查，已配置访问控制策略。", "符合", 1.0, "A-01"])

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_final_assessment_excel() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "出口防火墙-A"
    worksheet.append(["IP：143.8.51.2", "版本：V8.0.26"])
    worksheet.append(["扩展标准", "控制点", "测评项", "结果记录", "符合情况", "权重", "测评项编号"])
    worksheet.append(["安全通信网络", "边界访问控制", "应限制非授权访问", "经现场核查，已配置访问控制策略。", "符合", 1.0, "A-01"])
    worksheet.append([None, None, "应保留访问控制日志", "查看日志审计配置，当前已开启日志留存。", "部分符合", 0.6, "A-02"])

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
