from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.core.config import settings
from tests.history_excel_utils import build_final_assessment_excel


@pytest.fixture(autouse=True)
def use_mock_ocr_provider():
    original = settings.OCR_PROVIDER
    settings.OCR_PROVIDER = "mock"
    yield
    settings.OCR_PROVIDER = original


def create_project(client):
    resp = client.post(
        "/api/v1/projects",
        json={"code": "PJT-EXP", "name": "导出项目", "project_type": "等级保护测评", "status": "draft"},
    )
    assert resp.status_code == 201
    return resp.json()["data"]["id"]


def upload_evidence(client, project_id: str, filename: str, title: str, device: str) -> str:
    upload_resp = client.post(
        f"/api/v1/projects/{project_id}/evidences/upload",
        files={"file": (filename, f"hello export {filename}".encode("utf-8"), "text/plain")},
        data={
            "title": title,
            "evidence_type": "screenshot",
            "summary": "导出测试",
            "device": device,
            "tags_json": '["export"]',
            "source_ref": "test-export",
        },
    )
    assert upload_resp.status_code == 201
    return upload_resp.json()["data"]["id"]


def run_extract_flow(client, evidence_id: str, sample_id: str, template_code: str):
    ocr_resp = client.post(f"/api/v1/evidences/{evidence_id}/ocr", json={"sample_id": sample_id})
    assert ocr_resp.status_code == 200
    extract_resp = client.post(
        f"/api/v1/evidences/{evidence_id}/extract-fields",
        json={"template_code": template_code},
    )
    assert extract_resp.status_code == 200


def generate_record(client, project_id: str, evidence_id: str):
    resp = client.post(
        f"/api/v1/projects/{project_id}/records/generate",
        json={"evidence_id": evidence_id},
    )
    assert resp.status_code == 201
    return resp.json()["data"]


def import_project_template(client, project_id: str, filename: str = "reference.xlsx"):
    resp = client.post(
        f"/api/v1/projects/{project_id}/templates/import-reference",
        files={"file": (filename, build_final_assessment_excel(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 201
    return resp.json()["data"]


def advance_record_to_approved(client, record_id: str, current_status: str, reviewer: str):
    if current_status == "generated":
        review_resp = client.post(
            f"/api/v1/records/{record_id}/review",
            json={
                "status": "reviewed",
                "final_content": f"{record_id}-reviewed",
                "review_comment": "进入复核",
                "reviewed_by": reviewer,
            },
        )
        assert review_resp.status_code == 200
        current_status = "reviewed"

    if current_status == "reviewed":
        approve_resp = client.post(
            f"/api/v1/records/{record_id}/review",
            json={
                "status": "approved",
                "final_content": f"{record_id}-approved",
                "review_comment": "审批通过",
                "reviewed_by": reviewer,
            },
        )
        assert approve_resp.status_code == 200
        current_status = "approved"

    return current_status


def prepare_guidance_and_history(client, tmp_path):
    from tests.test_guidance_api import import_network_guidance, import_sample_history

    guidance_resp = import_network_guidance(client, tmp_path)
    history_resp = import_sample_history(client)
    assert guidance_resp.status_code == 201
    assert history_resp.status_code == 201
    guidance_id = client.get("/api/v1/guidance/items", params={"keyword": "边界防护"}).json()["data"]["items"][0]["id"]
    link_resp = client.post(f"/api/v1/guidance/{guidance_id}/link-history")
    assert link_resp.status_code == 200
    return guidance_id


def confirm_asset_and_guidance(client, project_id: str, evidence_id: str, asset_name: str, primary_ip: str, guidance_id: str) -> None:
    asset_resp = client.post(
        f"/api/v1/projects/{project_id}/assets",
        json={
            "filename": asset_name,
            "category": "server",
            "category_label": "服务器",
            "relative_path": f"assets/{asset_name}.txt",
            "primary_ip": primary_ip,
        },
    )
    assert asset_resp.status_code == 201
    asset_id = asset_resp.json()["data"]["id"]

    confirm_asset_resp = client.post(f"/api/v1/evidences/{evidence_id}/confirm-asset", json={"asset_id": asset_id})
    assert confirm_asset_resp.status_code == 200

    confirm_guidance_resp = client.post(f"/api/v1/evidences/{evidence_id}/confirm-guidance", json={"guidance_id": guidance_id})
    assert confirm_guidance_resp.status_code == 200


def open_workbook_from_response(response):
    return load_workbook(BytesIO(response.content))


def header_values(worksheet):
    return [cell.value for cell in worksheet[1]]


def test_project_export_txt_grouped_by_device(client):
    project_id = create_project(client)
    evidence_id_1 = upload_evidence(client, project_id, "firewall_basic.txt", "防火墙配置", "服务器A")
    evidence_id_2 = upload_evidence(client, project_id, "policy_missing_action.txt", "安全策略", "服务器A")
    evidence_id_3 = upload_evidence(client, project_id, "windows_host_partial.txt", "主机检查", "交换机B")

    run_extract_flow(client, evidence_id_1, "firewall_basic", "security_device_basic")
    run_extract_flow(client, evidence_id_2, "security_policy_missing_action", "security_policy_basic")
    run_extract_flow(client, evidence_id_3, "windows_host_partial", "host_basic_info")

    record_1 = generate_record(client, project_id, evidence_id_1)
    record_2 = generate_record(client, project_id, evidence_id_2)
    record_3 = generate_record(client, project_id, evidence_id_3)

    assert advance_record_to_approved(client, record_1["id"], record_1["status"], "alice") == "approved"
    assert advance_record_to_approved(client, record_2["id"], record_2["status"], "bob") == "approved"
    assert advance_record_to_approved(client, record_3["id"], record_3["status"], "carol") == "approved"

    export_resp = client.post(f"/api/v1/projects/{project_id}/export", json={"format": "txt"})
    assert export_resp.status_code == 201
    export_body = export_resp.json()
    assert export_body["message"] == "项目导出成功"
    export_id = export_body["data"]["id"]
    assert export_body["data"]["record_count"] == 3
    assert export_body["data"]["format"] == "txt"

    jobs_resp = client.get(f"/api/v1/projects/{project_id}/export-jobs")
    assert jobs_resp.status_code == 200
    jobs_body = jobs_resp.json()
    assert jobs_body["meta"]["total"] == 1
    assert jobs_body["data"][0]["id"] == export_id

    download_resp = client.get(f"/api/v1/exports/{export_id}/download")
    assert download_resp.status_code == 200
    assert download_resp.headers["content-type"].startswith("text/plain")
    content = download_resp.text
    assert "项目：导出项目" in content
    assert "设备：服务器A" in content
    assert "设备：交换机B" in content
    assert record_1["title"] in content
    assert record_2["title"] in content
    assert record_3["title"] in content
    assert "状态：exported" in content


def test_project_export_excel_official_multi_sheet(client, tmp_path):
    guidance_id = prepare_guidance_and_history(client, tmp_path)
    project_id = create_project(client)

    evidence_id_1 = upload_evidence(client, project_id, "firewall_basic.txt", "防火墙配置", "设备A")
    evidence_id_2 = upload_evidence(client, project_id, "policy_missing_action.txt", "安全策略", "设备B")

    run_extract_flow(client, evidence_id_1, "firewall_basic", "security_device_basic")
    run_extract_flow(client, evidence_id_2, "security_policy_missing_action", "security_policy_basic")
    confirm_asset_and_guidance(client, project_id, evidence_id_1, "FW-01", "10.0.0.1", guidance_id)
    confirm_asset_and_guidance(client, project_id, evidence_id_2, "SW-01", "10.0.0.2", guidance_id)

    record_1 = generate_record(client, project_id, evidence_id_1)
    record_2 = generate_record(client, project_id, evidence_id_2)
    assert advance_record_to_approved(client, record_1["id"], record_1["status"], "alice") == "approved"
    assert advance_record_to_approved(client, record_2["id"], record_2["status"], "bob") == "approved"

    export_resp = client.post(f"/api/v1/projects/{project_id}/export-excel", json={"mode": "official"})
    assert export_resp.status_code == 201
    body = export_resp.json()
    assert body["data"]["format"] == "xlsx"
    assert body["data"]["mode"] == "official"
    export_id = body["data"]["id"]

    download_resp = client.get(f"/api/v1/exports/{export_id}/download")
    assert download_resp.status_code == 200
    assert download_resp.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    workbook = open_workbook_from_response(download_resp)
    assert set(workbook.sheetnames) == {"FW-01", "SW-01"}
    worksheet = workbook["FW-01"]
    assert header_values(worksheet) == ["编号", "扩展标准", "控制点", "测评项", "结果记录", "符合情况", "分值"]
    assert worksheet.freeze_panes == "A2"
    row = [worksheet.cell(2, index).value for index in range(1, 8)]
    assert row[0] == "A-01"
    assert row[1] == "安全通信网络"
    assert row[5] == record_1["conclusion"]
    assert row[5] == "待人工确认"
    assert row[6] == 1


def test_project_export_excel_debug_contains_debug_columns(client, tmp_path):
    guidance_id = prepare_guidance_and_history(client, tmp_path)
    project_id = create_project(client)
    evidence_id = upload_evidence(client, project_id, "firewall_basic.txt", "防火墙配置", "设备A")

    run_extract_flow(client, evidence_id, "firewall_basic", "security_device_basic")
    confirm_asset_and_guidance(client, project_id, evidence_id, "FW-01", "10.0.0.1", guidance_id)

    record = generate_record(client, project_id, evidence_id)
    assert advance_record_to_approved(client, record["id"], record["status"], "alice") == "approved"

    export_resp = client.post(f"/api/v1/projects/{project_id}/export-excel", json={"mode": "debug"})
    assert export_resp.status_code == 201
    export_id = export_resp.json()["data"]["id"]

    download_resp = client.get(f"/api/v1/exports/{export_id}/download")
    workbook = open_workbook_from_response(download_resp)
    worksheet = workbook["FW-01"]
    headers = header_values(worksheet)
    assert headers == ["编号", "扩展标准", "控制点", "测评项", "结果记录", "符合情况", "分值", "测试对象", "证据文件", "指导书依据", "匹配分数", "历史样本ID"]
    row = [worksheet.cell(2, index).value for index in range(1, 13)]
    assert row[7] == "FW-01"
    assert row[8] == "防火墙配置"
    assert "边界防护" in row[9]
    assert "记录:" in row[10]
    assert row[11]


def test_project_export_excel_prefers_project_template_snapshot(client):
    project_id = create_project(client)
    import_project_template(client, project_id)
    evidence_id = upload_evidence(client, project_id, "firewall_basic.txt", "防火墙配置", "设备A")
    run_extract_flow(client, evidence_id, "firewall_basic", "security_device_basic")

    record = generate_record(client, project_id, evidence_id)
    approve_resp = client.post(
        f"/api/v1/records/{record['id']}/review",
        json={
            "status": "reviewed",
            "final_content": "人工复核后的模板记录",
            "review_comment": "进入复核",
            "reviewed_by": "alice",
        },
    )
    assert approve_resp.status_code == 200
    approve_resp = client.post(
        f"/api/v1/records/{record['id']}/review",
        json={
            "status": "approved",
            "final_content": "人工复核后的模板记录",
            "review_comment": "审批通过",
            "reviewed_by": "alice",
        },
    )
    assert approve_resp.status_code == 200

    export_resp = client.post(f"/api/v1/projects/{project_id}/export-excel", json={"mode": "official"})
    assert export_resp.status_code == 201
    export_id = export_resp.json()["data"]["id"]

    download_resp = client.get(f"/api/v1/exports/{export_id}/download")
    workbook = open_workbook_from_response(download_resp)
    assert workbook.sheetnames == ["出口防火墙-A"]
    worksheet = workbook["出口防火墙-A"]
    row = [worksheet.cell(2, index).value for index in range(1, 8)]
    assert row == ["A-01", "安全通信网络", "边界访问控制", "应限制非授权访问", "人工复核后的模板记录", "符合", 1]


def test_project_export_requires_approved_records(client):
    project_id = create_project(client)
    evidence_id = upload_evidence(client, project_id, "firewall_basic.txt", "防火墙配置", "服务器A")
    run_extract_flow(client, evidence_id, "firewall_basic", "security_device_basic")
    record = generate_record(client, project_id, evidence_id)
    assert record["status"] in {"generated", "reviewed"}

    export_resp = client.post(f"/api/v1/projects/{project_id}/export", json={"format": "txt"})
    assert export_resp.status_code == 400
    assert export_resp.json()["error"]["code"] == "RECORD_NOT_READY_FOR_EXPORT"


def test_project_export_unsupported_format(client):
    project_id = create_project(client)

    export_resp = client.post(f"/api/v1/projects/{project_id}/export", json={"format": "pdf"})
    assert export_resp.status_code == 400
    assert export_resp.json()["error"]["code"] == "EXPORT_FORMAT_NOT_SUPPORTED"
