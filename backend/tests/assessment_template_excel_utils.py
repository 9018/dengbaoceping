from io import BytesIO

from openpyxl import Workbook


def build_assessment_template_excel() -> bytes:
    workbook = Workbook()
    firewall = workbook.active
    firewall.title = "外联防火墙A"
    firewall.append(["设备：外联防火墙A", "版本：V1.0"])
    firewall.append(["扩展标准", "控制点", "测评项", "结果记录", "符合情况", "权重", "编号", "样例1", "样例2", "样例3"])
    firewall.append(["安全通信网络", "身份鉴别", "a）应对管理员身份进行鉴别", "经现场核查，管理员账号已启用口令认证。", "符合", 1.0, "A-01", "样例A1", "样例A2", "样例A3"])
    firewall.append([None, None, None, None, None, None, None, None, None, None])
    firewall.append([None, "访问控制", "a）应限制非授权访问", "经现场核查，已配置访问控制策略。", "部分符合", 0.7, "A-02"])

    switch = workbook.create_sheet("核心交换机")
    switch.append(["扩展标准", "控制点", "测评项", "结果记录", "符合情况", "分值", "测评项编号"])
    switch.append(["安全区域边界", "安全审计", "b）应记录关键操作", "当前已开启日志审计并转储。", "符合", 0.8, "B-01"])

    windows = workbook.create_sheet("Windows服务器")
    windows.append(["扩展标准", "控制点", "测评项", "结果记录", "符合情况", "分值", "编号"])
    windows.append(["安全计算环境", "访问控制", "a）应配置账户权限", "已按角色分配账户权限。", "符合", 0.6, "C-01"])

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_assessment_template_match_excel() -> bytes:
    workbook = Workbook()
    firewall = workbook.active
    firewall.title = "外联防火墙A"
    firewall.append(["设备：外联防火墙A", "版本：V2.0"])
    firewall.append(["扩展标准", "控制点", "测评项", "结果记录", "符合情况", "权重", "编号", "样例1"])
    firewall.append([
        "安全通信网络",
        "身份鉴别",
        "a）应对管理员身份进行鉴别",
        "查看系统-管理员账号-密码安全策略，确认最小密码长度、复杂度、有效期和口令认证已启用。",
        "符合",
        1.0,
        "A-01",
        "display password policy",
    ])
    firewall.append([
        None,
        "安全审计",
        "a）应记录管理员操作日志",
        "查看监控-日志页面，确认审计日志、操作日志和管理员操作记录已开启并留存。",
        "符合",
        0.9,
        "A-03",
        "日志页面截图",
    ])

    switch = workbook.create_sheet("核心交换机")
    switch.append(["扩展标准", "控制点", "测评项", "结果记录", "符合情况", "分值", "测评项编号", "样例1"])
    switch.append([
        "安全区域边界",
        "身份鉴别",
        "b）应限制登录失败次数并设置口令策略",
        "执行 display password-control，查看登录失败锁定阈值、最小密码长度和复杂度配置。",
        "符合",
        0.8,
        "B-02",
        "display password-control",
    ])
    switch.append([
        None,
        "安全审计",
        "b）应记录关键操作",
        "当前已开启日志审计并转储。",
        "符合",
        0.7,
        "B-01",
        "audit log",
    ])

    windows = workbook.create_sheet("Windows服务器")
    windows.append(["扩展标准", "控制点", "测评项", "结果记录", "符合情况", "分值", "编号", "样例1"])
    windows.append([
        "安全计算环境",
        "身份鉴别",
        "a）应配置本地密码策略和账户锁定策略",
        "打开本地安全策略 secpol.msc，查看密码策略中的最小密码长度、密码复杂度和账户锁定阈值。",
        "符合",
        0.85,
        "C-02",
        "本地安全策略",
    ])
    windows.append([
        None,
        "访问控制",
        "a）应配置账户权限",
        "已按角色分配账户权限。",
        "符合",
        0.6,
        "C-01",
        "账户权限",
    ])

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_assessment_template_excel_without_header() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "无效模板"
    worksheet.append(["说明", "这里没有合法表头"])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
